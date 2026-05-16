from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def export_excel(data, columns, title):

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # =========================
    # 🔹 TÍTULO
    # =========================
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(columns)
    )

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # =========================
    # 🔹 HEADERS
    # =========================
    for col_num, (header, key) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # =========================
    # 🔹 DATA
    # =========================
    for row_num, row in enumerate(data, start=4):
        for col_num, (header, key) in enumerate(columns, start=1):

            value = row.get(key)

            # Formatear fechas automáticamente
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")

            ws.cell(row=row_num, column=col_num, value=value)

    # =========================
    # 🔹 AUTO AJUSTE COLUMNAS (FIX ERROR)
    # =========================
    MAX_WIDTH = 40  # 👈 ancho máximo permitido
    for i, column_cells in enumerate(ws.iter_cols(min_row=3), 1):
        max_length = 0

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, MAX_WIDTH)
        col_letter = get_column_letter(i)

        ws.column_dimensions[col_letter].width = adjusted_width

    # =========================
    # 🔹 CREAR CARPETA
    # =========================
    os.makedirs("reports", exist_ok=True)

    # =========================
    # 🔹 GUARDAR ARCHIVO
    # =========================
    filename = f"reports/reporte_{int(datetime.now().timestamp())}.xlsx"

    wb.save(filename)

    return filename
    

def export_estadisticas_excel(data: dict, title: str) -> str:
    """
    Exportador específico para el reporte de estadísticas.
    Escribe 3 secciones en una misma hoja con el mismo formato
    que el reporte manual.
    """
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas"
 
    # ── Estilos ───────────────────────────────────────────────────────────────
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    center = Alignment(horizontal="center")
 
    def write_title(row, text, ncols):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12)
        cell.alignment = center
 
    def write_header(row, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
 
    def write_row(row, values):
        for col, v in enumerate(values, 1):
            ws.cell(row=row, column=col, value=v)
 
    current_row = 1
 
    # ── Título principal ──────────────────────────────────────────────────────
    write_title(current_row, title, 6)
    current_row += 2
 
    # ═════════════════════════════════════════════════════════════════════════
    # TABLA 1: Por semestre
    # ═════════════════════════════════════════════════════════════════════════
    write_title(current_row, "Resumen por semestre", 6)
    current_row += 1
 
    headers1 = [
        "Semestre",
        "Proyectos recibidos",
        "Proyectos Vigentes",
        "No. Estudiantes Proyectos Vigentes",
    ]
    write_header(current_row, headers1)
    current_row += 1
 
    total_recibidos = total_vigentes = total_estudiantes = 0
 
    for r in data["por_semestre"]:
        tec = r.get("vigentes_tecnologico", 0) or 0
        pro = r.get("vigentes_profesional", 0) or 0
        vigentes = r.get("proyectos_vigentes", 0) or 0
        vigentes_str = f"{vigentes} ({tec} Tecnología y {pro} Ingeniería)"
 
        write_row(current_row, [
            r["periodo"],
            r.get("proyectos_recibidos", 0),
            vigentes_str,
            r.get("estudiantes_vigentes", 0),
        ])
        total_recibidos  += r.get("proyectos_recibidos", 0) or 0
        total_vigentes   += vigentes
        total_estudiantes += r.get("estudiantes_vigentes", 0) or 0
        current_row += 1
 
    # Fila TOTAL
    total_row = current_row
    write_row(current_row, ["TOTAL", total_recibidos, total_vigentes, total_estudiantes])
    for col in range(1, 5):
        ws.cell(row=total_row, column=col).font = bold
    current_row += 2
 
    # ═════════════════════════════════════════════════════════════════════════
    # TABLA 2: Aprobación
    # ═════════════════════════════════════════════════════════════════════════
    write_title(current_row, "Aprobación de propuestas por semestre", 6)
    current_row += 1
 
    headers2 = [
        "Semestre",
        "Proyectos recibidos",
        "Propuestas aprobadas",
        "Propuestas pendientes por aprobar",
    ]
    write_header(current_row, headers2)
    current_row += 1
 
    for r in data["aprobacion"]:
        rec_est  = r.get("recibidos_estudiantes", 0) or 0
        apr_est  = r.get("aprobados_estudiantes", 0) or 0
        pen_est  = r.get("pendientes_estudiantes", 0) or 0
        apr      = r.get("aprobados", 0) or 0
        pen      = r.get("pendientes", 0) or 0
        rec      = r.get("recibidos", 0) or 0
 
        write_row(current_row, [
            r["periodo"],
            f"{rec} ({rec_est} estudiantes)",
            f"{apr} (Corresponde a {apr_est} estudiantes)",
            f"{pen} (Corresponde a {pen_est} estudiantes)",
        ])
        current_row += 1
 
    current_row += 1
 
    # ═════════════════════════════════════════════════════════════════════════
    # TABLA 3: Por modalidad
    # ═════════════════════════════════════════════════════════════════════════
    write_title(current_row, "Distribución por modalidad y programa", 8)
    current_row += 1
 
    # Sub-encabezados de programa
    ws.cell(row=current_row, column=2,
            value="Tecnología en Desarrollo de Sistemas Informáticos").font = bold
    ws.cell(row=current_row, column=4,
            value="Ingeniería de Sistemas").font = bold
    ws.merge_cells(start_row=current_row, start_column=2,
                   end_row=current_row, end_column=3)
    ws.merge_cells(start_row=current_row, start_column=4,
                   end_row=current_row, end_column=5)
    current_row += 1
 
    headers3 = [
        "Modalidad",
        "Cantidad Proyectos", "No. Estudiantes",
        "Cantidad Proyectos", "No. Estudiantes",
    ]
    write_header(current_row, headers3)
    current_row += 1
 
    total_pt = total_et = total_pp = total_ep = 0
 
    for r in data["por_modalidad"]:
        pt = r.get("proyectos_tecnologico", 0) or 0
        et = r.get("estudiantes_tecnologico", 0) or 0
        pp = r.get("proyectos_profesional", 0) or 0
        ep = r.get("estudiantes_profesional", 0) or 0
 
        write_row(current_row, [r.get("modalidad", ""), pt, et, pp, ep])
        total_pt += pt; total_et += et
        total_pp += pp; total_ep += ep
        current_row += 1
 
    # Fila Total
    write_row(current_row, ["Total", total_pt, total_et, total_pp, total_ep])
    for col in range(1, 6):
        ws.cell(row=current_row, column=col).font = bold
    current_row += 1
 
    # ── Auto ajuste columnas ──────────────────────────────────────────────────
    for i, col_cells in enumerate(ws.iter_cols(), 1):
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value), default=10
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)
 
    # ── Guardar ───────────────────────────────────────────────────────────────
    os.makedirs("reports", exist_ok=True)
    filename = f"reports/reporte_{int(datetime.now().timestamp())}.xlsx"
    wb.save(filename)
    return filename
 
def export_dinamica_profesores_excel(data: dict, title: str) -> str:
    """
    Exportador para dinámica de profesores.
    Dos tablas pivote lado a lado:
      Izquierda: Directores × Periodo
      Derecha:   Evaluadores × Periodo
    """
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Dinámica Profesores"
 
    periodos   = data["periodos"]
    directores = data["directores"]
    evaluadores = data["evaluadores"]
 
    bold        = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    center      = Alignment(horizontal="center")
 
    # ── Columnas ──────────────────────────────────────────────────────────────
    # Tabla directores: col 1=nombre, 2..n=periodos, n+1=Total
    # Separador: col n+2
    # Tabla evaluadores: col n+3=nombre, n+4..=periodos, last=Total
    n = len(periodos)
    sep_col   = n + 2          # columna separadora (vacía)
    eva_start = n + 3          # primera col de evaluadores
 
    # ── Fila 1: título principal ──────────────────────────────────────────────
    total_cols = eva_start + n + 1
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=total_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=13)
    c.alignment = center
 
    # ── Fila 2: subtítulos de sección ─────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2,   end_column=n + 1)
    c = ws.cell(row=2, column=1,
                value="ASIGNACIÓN DE DIRECCIÓN DE TRABAJO DE GRADO")
    c.font = bold; c.alignment = center
 
    ws.merge_cells(start_row=2, start_column=eva_start,
                   end_row=2,   end_column=eva_start + n)
    c = ws.cell(row=2, column=eva_start,
                value="ASIGNACIÓN DE EVALUADOR DE TRABAJO DE GRADO")
    c.font = bold; c.alignment = center
 
    # ── Fila 3: subtítulo aprobación ──────────────────────────────────────────
    ws.merge_cells(start_row=3, start_column=1,
                   end_row=3,   end_column=n + 1)
    c = ws.cell(row=3, column=1,
                value="APROBACIÓN PROPUESTA F-DC-125 - F-DC-128")
    c.font = bold; c.alignment = center
 
    ws.merge_cells(start_row=3, start_column=eva_start,
                   end_row=3,   end_column=eva_start + n)
    c = ws.cell(row=3, column=eva_start,
                value="APROBACIÓN PROPUESTA F-DC-125 - F-DC-128")
    c.font = bold; c.alignment = center
 
    # ── Fila 4: encabezado "Proyectos / PERIODO" ──────────────────────────────
    for base_col, label in [(1, "DIRECTOR"), (eva_start, "EVALUADOR")]:
        c = ws.cell(row=4, column=base_col, value=label)
        c.font = bold; c.fill = header_fill; c.alignment = center
        for i, p in enumerate(periodos):
            col = base_col + 1 + i
            c = ws.cell(row=4, column=col, value=p)
            c.font = bold; c.fill = header_fill; c.alignment = center
        c = ws.cell(row=4, column=base_col + n + 1, value="Total general")
        c.font = bold; c.fill = header_fill; c.alignment = center
 
    # ── Filas de datos: directores ────────────────────────────────────────────
    dir_names = sorted(directores.keys())
    eva_names = sorted(evaluadores.keys())
    max_rows  = max(len(dir_names), len(eva_names))
 
    dir_totals = [0] * (n + 1)   # totales por columna [periodos..., grand_total]
    eva_totals = [0] * (n + 1)
 
    for row_i in range(max_rows):
        row = 5 + row_i
 
        # Director
        if row_i < len(dir_names):
            name = dir_names[row_i]
            ws.cell(row=row, column=1, value=name)
            total = 0
            for j, p in enumerate(periodos):
                val = directores[name].get(p, None)
                ws.cell(row=row, column=2 + j, value=val)
                if val:
                    total += val
                    dir_totals[j] += val
            ws.cell(row=row, column=2 + n, value=total)
            dir_totals[n] += total
 
        # Evaluador
        if row_i < len(eva_names):
            name = eva_names[row_i]
            ws.cell(row=row, column=eva_start, value=name)
            total = 0
            for j, p in enumerate(periodos):
                val = evaluadores[name].get(p, None)
                ws.cell(row=row, column=eva_start + 1 + j, value=val)
                if val:
                    total += val
                    eva_totals[j] += val
            ws.cell(row=row, column=eva_start + 1 + n, value=total)
            eva_totals[n] += total
 
    # ── Fila de totales ───────────────────────────────────────────────────────
    total_row = 5 + max_rows
    ws.cell(row=total_row, column=1, value="Total general").font = bold
    for j in range(n + 1):
        c = ws.cell(row=total_row, column=2 + j, value=dir_totals[j] or None)
        c.font = bold
    ws.cell(row=total_row, column=eva_start, value="Total general").font = bold
    for j in range(n + 1):
        c = ws.cell(row=total_row, column=eva_start + 1 + j,
                    value=eva_totals[j] or None)
        c.font = bold
 
    # ── Auto ajuste columnas ──────────────────────────────────────────────────
    for i, col_cells in enumerate(ws.iter_cols(), 1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=4)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)
 
    # ── Guardar ───────────────────────────────────────────────────────────────
    os.makedirs("reports", exist_ok=True)
    filename = f"reports/reporte_{int(datetime.now().timestamp())}.xlsx"
    wb.save(filename)
    return filename

def export_dinamica_modalidad_excel(data: dict, title: str) -> str:
    """
    Exportador dinámica por modalidad.
    Layout jerárquico: fila de periodo (bold) + filas de modalidad debajo.
    """
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Dinámica x Modalidad"
 
    bold        = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    period_fill = PatternFill("solid", fgColor="F2F2F2")
    center      = Alignment(horizontal="center")
 
    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=13)
    c.alignment = center
 
    # ── Encabezados ───────────────────────────────────────────────────────────
    headers = ["MODALIDADES", "Cantidad por Modalidad", "APROBADOS POR MODALIDAD"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = center
 
    current_row = 3
    grand_total_cantidad = 0
    grand_total_aprobados = 0
 
    for periodo in sorted(data.keys()):
        filas = data[periodo]
        total_periodo    = sum(f["cantidad"]  for f in filas)
        aprobados_periodo = sum(f["aprobados"] for f in filas)
 
        # ── Fila de periodo (agrupación) ──────────────────────────────────────
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).fill = period_fill
        ws.cell(row=current_row, column=1, value=periodo).font        = bold
        ws.cell(row=current_row, column=2, value=total_periodo).font  = bold
        ws.cell(row=current_row, column=3, value=aprobados_periodo if aprobados_periodo else None).font = bold
        current_row += 1
 
        # ── Filas de modalidad ────────────────────────────────────────────────
        for f in filas:
            ws.cell(row=current_row, column=1, value=f["modalidad"])
            ws.cell(row=current_row, column=2, value=f["cantidad"])
            ws.cell(row=current_row, column=3,
                    value=f["aprobados"] if f["aprobados"] else None)
            current_row += 1
 
        grand_total_cantidad  += total_periodo
        grand_total_aprobados += aprobados_periodo
 
    # ── Fila Total general ────────────────────────────────────────────────────
    ws.cell(row=current_row, column=1, value="Total general").font = bold
    ws.cell(row=current_row, column=2, value=grand_total_cantidad).font = bold
    ws.cell(row=current_row, column=3, value=grand_total_aprobados).font = bold
 
    # ── Auto ajuste columnas ──────────────────────────────────────────────────
    for i, col_cells in enumerate(ws.iter_cols(), 1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 45)
 
    # ── Guardar ───────────────────────────────────────────────────────────────
    os.makedirs("reports", exist_ok=True)
    filename = f"reports/reporte_{int(datetime.now().timestamp())}.xlsx"
    wb.save(filename)
    return filename


